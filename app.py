import streamlit as st
import openpyxl
import json
import google.generativeai as genai
import os
from datetime import datetime

# --- CONFIGURACIÓN ---
try:
    API_KEY = st.secrets["GEMINI_API_KEY"]
except:
    API_KEY = "TU_CLAVE_LOCAL_AQUI" 

genai.configure(api_key=API_KEY)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(BASE_DIR, "CRD_Tesis_Cardiorrenal.xlsx")

model = genai.GenerativeModel('gemini-1.5-flash')

st.set_page_config(page_title="CRD Tesis - Ingreso de Datos", layout="wide")
st.title("🩺 Extracción Inteligente y Formulario (V2)")

# --- DICCIONARIO DE VARIABLES ---
claves_por_defecto = {
    "id_pac": "", "fecha_inc": datetime.today().strftime('%d/%m/%Y'), "edad": "", "sexo": "H", 
    "peso": "", "talla": "", "eti_erc": "HTA", "eti_ic": "HFpEF", 
    "dm2": "No", "hta": "No", "fa": "No", "epoc": "No", "sd_metab": "No",
    "tabaco": "No", "enolismo": "No", "hepato": "No",
    "hb": "", "creat": "", "cist_c": "", "fge": "", "urea": "", "ac_urico": "", "prot_creat": "",
    "ast": "", "alt": "", "plaq": "", "bili_t": "", "bili_d": "", "albumina": "", "hba1c": "", "colest": "",
    "nt_probnp": "", "ca125": "", "gal3": "", "sst2": "", "gdf15": "", "biobanco": "No",
    "fevi": "", "gls": "", "masa_vi": "", "tapse": "", "vai": "", "vci": "",
    "lsm": "", "cat_fibro": "F0-F1 (<7)", "cap": "", "med_val": "", "iqr_med": "Sí", 
    "dias_desc": "", "nt_prueba": "", "edemas_prueba": "No",
    "ieca": "No", "ara2": "No", "bb": "No", "amr": "No", "sac_val": "No", "sglt2i": "No",
    "diur_asa": "No", "hctz": "No", "acetazolamida": "No", "estatinas": "No", "epo": "No",
    "meses_seg": "", "m_cv": "No", "f_m_cv": "", "hosp_ic": "No", "f_hosp_ic": "", 
    "iam": "No", "f_iam": "", "acv": "No", "f_acv": "", "m_tot": "No", "f_m_tot": "",
    "trs": "No", "f_trs": "", "caida_fge": "No", "f_caida_fge": "", "sd_cr": "No", "f_sd_cr": ""
}

# Inicializar estado
for clave, valor in claves_por_defecto.items():
    if clave not in st.session_state:
        st.session_state[clave] = valor

# --- FUNCIONES DE BASE DE DATOS ---
def cargar_paciente(id_buscado):
    if not os.path.exists(EXCEL_FILE): return False
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws_clin = wb["01_Datos_Clinicos"]
    
    fila_obj = None
    for r in range(2, ws_clin.max_row + 1):
        if str(ws_clin.cell(row=r, column=1).value).strip() == str(id_buscado).strip():
            fila_obj = r
            break
            
    if not fila_obj: return False

    def get_val(ws_name, col):
        v = wb[ws_name].cell(row=fila_obj, column=col).value
        return "" if v is None else str(v).strip()

    # Mapeo de carga
    st.session_state.fecha_inc = get_val("01_Datos_Clinicos", 2)
    st.session_state.edad = get_val("01_Datos_Clinicos", 3)
    st.session_state.sexo = get_val("01_Datos_Clinicos", 4)
    st.session_state.peso = get_val("01_Datos_Clinicos", 5)
    st.session_state.talla = get_val("01_Datos_Clinicos", 6)
    st.session_state.eti_erc = get_val("01_Datos_Clinicos", 8)
    st.session_state.eti_ic = get_val("01_Datos_Clinicos", 9)
    st.session_state.dm2 = get_val("01_Datos_Clinicos", 10)
    st.session_state.hta = get_val("01_Datos_Clinicos", 11)
    st.session_state.fa = get_val("01_Datos_Clinicos", 12)
    st.session_state.epoc = get_val("01_Datos_Clinicos", 13)
    st.session_state.sd_metab = get_val("01_Datos_Clinicos", 14)
    st.session_state.tabaco = get_val("01_Datos_Clinicos", 15)
    st.session_state.enolismo = get_val("01_Datos_Clinicos", 16)
    st.session_state.hepato = get_val("01_Datos_Clinicos", 17)

    st.session_state.hb = get_val("02_Analitica_Biomarcadores", 2)
    st.session_state.creat = get_val("02_Analitica_Biomarcadores", 3)
    st.session_state.cist_c = get_val("02_Analitica_Biomarcadores", 4)
    st.session_state.fge = get_val("02_Analitica_Biomarcadores", 5)
    st.session_state.urea = get_val("02_Analitica_Biomarcadores", 6)
    st.session_state.ac_urico = get_val("02_Analitica_Biomarcadores", 7)
    st.session_state.prot_creat = get_val("02_Analitica_Biomarcadores", 8)
    st.session_state.ast = get_val("02_Analitica_Biomarcadores", 9)
    st.session_state.alt = get_val("02_Analitica_Biomarcadores", 10)
    st.session_state.plaq = get_val("02_Analitica_Biomarcadores", 11)
    st.session_state.bili_t = get_val("02_Analitica_Biomarcadores", 13)
    st.session_state.bili_d = get_val("02_Analitica_Biomarcadores", 14)
    st.session_state.albumina = get_val("02_Analitica_Biomarcadores", 15)
    st.session_state.hba1c = get_val("02_Analitica_Biomarcadores", 16)
    st.session_state.colest = get_val("02_Analitica_Biomarcadores", 17)
    st.session_state.nt_probnp = get_val("02_Analitica_Biomarcadores", 18)
    st.session_state.ca125 = get_val("02_Analitica_Biomarcadores", 19)
    st.session_state.gal3 = get_val("02_Analitica_Biomarcadores", 20)
    st.session_state.sst2 = get_val("02_Analitica_Biomarcadores", 21)
    st.session_state.gdf15 = get_val("02_Analitica_Biomarcadores", 22)
    st.session_state.biobanco = get_val("02_Analitica_Biomarcadores", 23)

    st.session_state.fevi = get_val("03_Eco_Elastografia", 2)
    st.session_state.gls = get_val("03_Eco_Elastografia", 3)
    st.session_state.masa_vi = get_val("03_Eco_Elastografia", 4)
    st.session_state.tapse = get_val("03_Eco_Elastografia", 5)
    st.session_state.vai = get_val("03_Eco_Elastografia", 6)
    st.session_state.vci = get_val("03_Eco_Elastografia", 7)
    st.session_state.lsm = get_val("03_Eco_Elastografia", 8)
    st.session_state.cat_fibro = get_val("03_Eco_Elastografia", 9)
    st.session_state.cap = get_val("03_Eco_Elastografia", 10)
    st.session_state.med_val = get_val("03_Eco_Elastografia", 11)
    st.session_state.iqr_med = get_val("03_Eco_Elastografia", 12)
    st.session_state.dias_desc = get_val("03_Eco_Elastografia", 13)
    st.session_state.nt_prueba = get_val("03_Eco_Elastografia", 14)
    st.session_state.edemas_prueba = get_val("03_Eco_Elastografia", 15)

    st.session_state.ieca = get_val("04_Tratamiento", 2)
    st.session_state.ara2 = get_val("04_Tratamiento", 3)
    st.session_state.bb = get_val("04_Tratamiento", 4)
    st.session_state.amr = get_val("04_Tratamiento", 5)
    st.session_state.sac_val = get_val("04_Tratamiento", 6)
    st.session_state.sglt2i = get_val("04_Tratamiento", 7)
    st.session_state.diur_asa = get_val("04_Tratamiento", 8)
    st.session_state.hctz = get_val("04_Tratamiento", 9)
    st.session_state.acetazolamida = get_val("04_Tratamiento", 10)
    st.session_state.estatinas = get_val("04_Tratamiento", 11)
    st.session_state.epo = get_val("04_Tratamiento", 12)

    st.session_state.meses_seg = get_val("05_Seguimiento_24m", 2)
    st.session_state.m_cv = get_val("05_Seguimiento_24m", 3)
    st.session_state.f_m_cv = get_val("05_Seguimiento_24m", 4)
    st.session_state.hosp_ic = get_val("05_Seguimiento_24m", 5)
    st.session_state.f_hosp_ic = get_val("05_Seguimiento_24m", 6)
    st.session_state.iam = get_val("05_Seguimiento_24m", 7)
    st.session_state.f_iam = get_val("05_Seguimiento_24m", 8)
    st.session_state.acv = get_val("05_Seguimiento_24m", 9)
    st.session_state.f_acv = get_val("05_Seguimiento_24m", 10)
    st.session_state.m_tot = get_val("05_Seguimiento_24m", 12)
    st.session_state.f_m_tot = get_val("05_Seguimiento_24m", 13)
    st.session_state.trs = get_val("05_Seguimiento_24m", 14)
    st.session_state.f_trs = get_val("05_Seguimiento_24m", 15)
    st.session_state.caida_fge = get_val("05_Seguimiento_24m", 16)
    st.session_state.f_caida_fge = get_val("05_Seguimiento_24m", 17)
    st.session_state.sd_cr = get_val("05_Seguimiento_24m", 18)
    st.session_state.f_sd_cr = get_val("05_Seguimiento_24m", 19)

    # Restaurar valores seguros por defecto para selects si están vacíos
    for k in claves_por_defecto.keys():
        if st.session_state[k] in ["None", "", None]:
            if claves_por_defecto[k] in ["Sí", "No", "H", "M", "HTA", "HFpEF", "F0-F1 (<7)"]:
                st.session_state[k] = claves_por_defecto[k]
                
    return True

# --- FUNCIÓN DE EXTRACCIÓN IA ---
def extraer_datos(texto):
    prompt = f"""
    Eres un asistente médico experto. Analiza el siguiente texto clínico y extrae todas las variables posibles para un estudio de insuficiencia cardíaca y renal.
    Devuelve ÚNICAMENTE un objeto JSON válido con las siguientes claves (si no encuentras el dato, pon null o dejalo vacío):
    - id_pac, edad, sexo, peso, talla, eti_erc, eti_ic
    - dm2, hta, fa, epoc, sd_metab, tabaco, enolismo, hepato (valores: Sí o No)
    - hb, creat, fge, ast, alt, plaq, nt_probnp, ca125, fevi, tapse, lsm
    - ieca, ara2, bb, amr, sac_val, sglt2i, diur_asa, hctz, acetazolamida (valores: Sí o No)

    Texto clínico a analizar:
    {texto}
    """
    try:
        res = model.generate_content(prompt)
        texto_json = res.text.strip().replace('```json', '').replace('```', '')
        return json.loads(texto_json)
    except Exception as e:
        return None

# --- MAQUETACIÓN DE LA INTERFAZ ---
col_izq, col_der = st.columns([1, 2])

with col_izq:
    st.subheader("1. Pegar Evolución / Historia")
    texto_input = st.text_area("Pega aquí los datos en texto plano:", height=300)
    
    if st.button("🧠 Auto-completar Formulario", type="primary", use_container_width=True):
        if texto_input:
            with st.spinner("Procesando historia clínica..."):
                datos_ia = extraer_datos(texto_input)
                if datos_ia:
                    for k, v in datos_ia.items():
                        if v is not None and str(v).strip() != "":
                            valor_str = str(v).strip()
                            if valor_str.lower() in ["si", "sí"]: valor_str = "Sí"
                            if valor_str.lower() in ["no"]: valor_str = "No"
                            if k in st.session_state:
                                st.session_state[k] = valor_str
                    st.success("Extracción completada. Revisa los campos.")
                    st.rerun()
                else:
                    st.error("No se pudo extraer la información.")
        else:
            st.warning("Pega un texto primero.")

with col_der:
    c_head1, c_head2 = st.columns([2, 1])
    with c_head1:
        st.subheader("2. Formulario de Datos")
    with c_head2:
        # Botón para limpiar toda la pantalla
        if st.button("🔄 Nuevo Paciente / Limpiar", use_container_width=True):
            for clave, valor in claves_por_defecto.items():
                st.session_state[clave] = valor
            st.rerun()
            
    tab1, tab2, tab3, tab4, tab5 = st.tabs(["01. Clínicos", "02. Analítica", "03. Eco/Fibro", "04. Tratamiento", "05. Seguimiento"])
    
    with tab1:
        # El bloque de ID y Búsqueda
        c1, c2, c3, c4 = st.columns([1.5, 0.7, 1, 1])
        c1.text_input("ID Paciente", key="id_pac")
        with c2:
            st.write("&nbsp;") 
            if st.button("🔍 Buscar", use_container_width=True):
                if st.session_state.id_pac:
                    if cargar_paciente(st.session_state.id_pac):
                        st.success("✅ Paciente cargado. Modo edición.")
                    else:
                        st.warning("⚠️ ID no existe. Se guardará como nuevo.")
        c3.text_input("Fecha Inclusión", key="fecha_inc")
        c4.selectbox("Sexo", ["H", "M"], key="sexo")
        
        c5, c6, c7, c8 = st.columns(4)
        c5.text_input("Edad", key="edad")
        c6.text_input("Peso (kg)", key="peso")
        c7.text_input("Talla (cm)", key="talla")
        c8.selectbox("Etiología ERC", ["DM2", "HTA", "Glomerulonefritis", "Poliquistosis", "Otras"], key="eti_erc")
        
        c_ic, _ = st.columns([1, 3])
        c_ic.selectbox("Etiología IC", ["Isquémica", "Hipertensiva", "MCD", "HFpEF", "Valvular"], key="eti_ic")
        
        st.markdown("**Comorbilidades Adicionales**")
        c9, c10, c11, c12 = st.columns(4)
        c9.selectbox("DM2", ["Sí", "No"], key="dm2")
        c10.selectbox("HTA", ["Sí", "No"], key="hta")
        c11.selectbox("FA", ["Sí", "No"], key="fa")
        c12.selectbox("EPOC", ["Sí", "No"], key="epoc")
        
        c13, c14, c15, c16 = st.columns(4)
        c13.selectbox("Sd. Metab", ["Sí", "No"], key="sd_metab")
        c14.selectbox("Tabaquismo", ["Sí", "No"], key="tabaco")
        c15.selectbox("Enolismo", ["Sí", "No"], key="enolismo")
        c16.selectbox("Hepatopatía", ["Sí", "No"], key="hepato")

    with tab2:
        c1, c2, c3, c4 = st.columns(4)
        c1.text_input("Hemoglobina", key="hb")
        c2.text_input("Creatinina", key="creat")
        c3.text_input("Cistatina C", key="cist_c")
        c4.text_input("FGe CKD-EPI", key="fge")
        
        c5, c6, c7, c8 = st.columns(4)
        c5.text_input("Prot/Creat", key="prot_creat")
        c6.text_input("AST", key="ast")
        c7.text_input("ALT", key="alt")
        c8.text_input("Plaquetas", key="plaq")
        
        st.markdown("**Biomarcadores y Muestras**")
        c9, c10, c11, c12 = st.columns(4)
        c9.text_input("NT-proBNP", key="nt_probnp")
        c10.text_input("CA125", key="ca125")
        c11.text_input("Galectina-3", key="gal3")
        c12.selectbox("Muestra a Biobanco", ["Sí", "No"], key="biobanco")

    with tab3:
        c1, c2, c3, c4 = st.columns(4)
        c1.text_input("FEVI (%)", key="fevi")
        c2.text_input("TAPSE", key="tapse")
        c3.text_input("LSM (kPa)", key="lsm")
        c4.text_input("CAP (dB/m)", key="cap")
        c5, c6 = st.columns(2)
        c5.text_input("NT-proBNP (día prueba)", key="nt_prueba")
        c6.selectbox("Edemas activos", ["Sí", "No"], key="edemas_prueba")

    with tab4:
        st.markdown("**Tratamiento Farmacológico Optimizado**")
        c1, c2, c3, c4 = st.columns(4)
        c1.selectbox("IECAS", ["Sí", "No"], key="ieca")
        c2.selectbox("ARA2", ["Sí", "No"], key="ara2")
        c3.selectbox("Betabloqueantes", ["Sí", "No"], key="bb")
        c4.selectbox("AMR", ["Sí", "No"], key="amr")
        
        c5, c6, c7, c8 = st.columns(4)
        c5.selectbox("SAC/VAL", ["Sí", "No"], key="sac_val")
        c6.selectbox("SGLT2i", ["Sí", "No"], key="sglt2i")
        c7.selectbox("Diurético de Asa", ["Sí", "No"], key="diur_asa")
        c8.selectbox("HCTZ", ["Sí", "No"], key="hctz")
        
        c9, c10, c11, _ = st.columns(4)
        c9.selectbox("Acetazolamida", ["Sí", "No"], key="acetazolamida")
        c10.selectbox("Estatinas", ["Sí", "No"], key="estatinas")
        c11.selectbox("Eritropoyetina", ["Sí", "No"], key="epo")

    with tab5:
        st.markdown("**Seguimiento a 24 Meses (Endpoints y Fechas)**")
        st.text_input("Meses de Seguimiento Total", key="meses_seg")
        st.markdown("---")
        
        e1, e2 = st.columns(2)
        e1.selectbox("Muerte Cardiovascular", ["Sí", "No"], key="m_cv")
        e2.text_input("Fecha Muerte CV", key="f_m_cv", placeholder="DD/MM/AAAA")
        
        e3, e4 = st.columns(2)
        e3.selectbox("Hosp. IC descompensada", ["Sí", "No"], key="hosp_ic")
        e4.text_input("Fecha Hosp. IC", key="f_hosp_ic", placeholder="DD/MM/AAAA")
        
        e5, e6 = st.columns(2)
        e5.selectbox("IAM no fatal", ["Sí", "No"], key="iam")
        e6.text_input("Fecha IAM", key="f_iam", placeholder="DD/MM/AAAA")
        
        e7, e8 = st.columns(2)
        e7.selectbox("ACV no fatal", ["Sí", "No"], key="acv")
        e8.text_input("Fecha ACV", key="f_acv", placeholder="DD/MM/AAAA")
        
        st.markdown("---")
        e9, e10 = st.columns(2)
        e9.selectbox("Muerte Total", ["Sí", "No"], key="m_tot")
        e10.text_input("Fecha Muerte Total", key="f_m_tot", placeholder="DD/MM/AAAA")
        
        e11, e12 = st.columns(2)
        e11.selectbox("Inicio TRS", ["Sí", "No"], key="trs")
        e12.text_input("Fecha TRS", key="f_trs", placeholder="DD/MM/AAAA")
        
        e13, e14 = st.columns(2)
        e13.selectbox("Caída FGe ≥25%", ["Sí", "No"], key="caida_fge")
        e14.text_input("Fecha Caída FGe", key="f_caida_fge", placeholder="DD/MM/AAAA")
        
        e15, e16 = st.columns(2)
        e15.selectbox("Sd. CR Agudo", ["Sí", "No"], key="sd_cr")
        e16.text_input("Fecha Sd. CR Agudo", key="f_sd_cr", placeholder="DD/MM/AAAA")

# --- BOTONES DE GUARDADO Y DESCARGA ---
st.markdown("---")

col_btn1, col_btn2 = st.columns(2)

with col_btn1:
    if st.button("💾 Guardar Registro en Excel", use_container_width=True):
        if not st.session_state.id_pac:
            st.error("⚠️ El ID de Paciente es obligatorio.")
        elif os.path.exists(EXCEL_FILE):
            try:
                wb = openpyxl.load_workbook(EXCEL_FILE)
                s = st.session_state 
                
                # Función que decide si sobrescribir o crear fila nueva
                def encontrar_fila(ws, id_buscado):
                    for r in range(2, ws.max_row + 2):
                        val = ws.cell(row=r, column=1).value
                        if val is not None and str(val).strip() == str(id_buscado).strip():
                            return r, True
                    for r in range(2, ws.max_row + 100):
                        val = ws.cell(row=r, column=1).value
                        if val is None or str(val).strip() == "":
                            return r, False
                    return 2, False

                # Usamos un sistema de inserción específico por columnas para NO borrar las celdas amarillas con fórmulas
                ws_clin = wb["01_Datos_Clinicos"]
                fila, existe = encontrar_fila(ws_clin, s.id_pac)
                datos_clin = [(1, s.id_pac), (2, s.fecha_inc), (3, s.edad), (4, s.sexo), (5, s.peso), (6, s.talla), 
                              (8, s.eti_erc), (9, s.eti_ic), (10, s.dm2), (11, s.hta), (12, s.fa), (13, s.epoc), 
                              (14, s.sd_metab), (15, s.tabaco), (16, s.enolismo), (17, s.hepato)]
                for col, val in datos_clin: ws_clin.cell(row=fila, column=col, value=val)
                
                ws_lab = wb["02_Analitica_Biomarcadores"]
                fila_lab, _ = encontrar_fila(ws_lab, s.id_pac)
                datos_lab = [(1, s.id_pac), (2, s.hb), (3, s.creat), (4, s.cist_c), (5, s.fge), (6, s.urea), 
                             (7, s.ac_urico), (8, s.prot_creat), (9, s.ast), (10, s.alt), (11, s.plaq), 
                             (13, s.bili_t), (14, s.bili_d), (15, s.albumina), (16, s.hba1c), (17, s.colest), 
                             (18, s.nt_probnp), (19, s.ca125), (20, s.gal3), (21, s.sst2), (22, s.gdf15), (23, s.biobanco)]
                for col, val in datos_lab: ws_lab.cell(row=fila_lab, column=col, value=val)
                
                ws_eco = wb["03_Eco_Elastografia"]
                fila_eco, _ = encontrar_fila(ws_eco, s.id_pac)
                datos_eco = [(1, s.id_pac), (2, s.fevi), (3, s.gls), (4, s.masa_vi), (5, s.tapse), (6, s.vai), (7, s.vci), 
                             (8, s.lsm), (9, s.cat_fibro), (10, s.cap), (11, s.med_val), (12, s.iqr_med), (13, s.dias_desc), 
                             (14, s.nt_prueba), (15, s.edemas_prueba)]
                for col, val in datos_eco: ws_eco.cell(row=fila_eco, column=col, value=val)
                
                ws_tx = wb["04_Tratamiento"]
                fila_tx, _ = encontrar_fila(ws_tx, s.id_pac)
                datos_tx = [(1, s.id_pac), (2, s.ieca), (3, s.ara2), (4, s.bb), (5, s.amr), (6, s.sac_val), (7, s.sglt2i), 
                            (8, s.diur_asa), (9, s.hctz), (10, s.acetazolamida), (11, s.estatinas), (12, s.epo)]
                for col, val in datos_tx: ws_tx.cell(row=fila_tx, column=col, value=val)

                ws_end = wb["05_Seguimiento_24m"]
                fila_end, _ = encontrar_fila(ws_end, s.id_pac)
                datos_end = [(1, s.id_pac), (2, s.meses_seg), (3, s.m_cv), (4, s.f_m_cv), (5, s.hosp_ic), (6, s.f_hosp_ic), 
                             (7, s.iam), (8, s.f_iam), (9, s.acv), (10, s.f_acv), (12, s.m_tot), (13, s.f_m_tot), 
                             (14, s.trs), (15, s.f_trs), (16, s.caida_fge), (17, s.f_caida_fge), (18, s.sd_cr), (19, s.f_sd_cr)]
                for col, val in datos_end: ws_end.cell(row=fila_end, column=col, value=val)
                
                wb.save(EXCEL_FILE)
                if existe:
                    st.success(f"✅ Paciente {s.id_pac} ACTUALIZADO correctamente.")
                else:
                    st.success(f"✅ Paciente {s.id_pac} NUEVO guardado correctamente.")
                
            except Exception as e:
                st.error(f"Error inesperado al guardar: {e}")
        else:
            st.error(f"No se encuentra {EXCEL_FILE}.")

with col_btn2:
    if os.path.exists(EXCEL_FILE):
        with open(EXCEL_FILE, "rb") as f:
            st.download_button(
                label="⬇️ Descargar Excel Actualizado",
                data=f,
                file_name="CRD_Tesis_Cardiorrenal_Actualizado.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
